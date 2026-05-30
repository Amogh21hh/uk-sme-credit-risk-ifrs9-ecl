"""Build a professional multi-tab IFRS 9 ECL workbook."""
import pandas as pd, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import BarChart, PieChart, Reference

OUT_XLSX = "/sessions/dreamy-vibrant-pascal/mnt/outputs/uk-sme-credit-risk/excel/UK_SME_ECL_Model.xlsx"
SCORED = "/sessions/dreamy-vibrant-pascal/mnt/outputs/uk-sme-credit-risk/outputs/scored_portfolio.csv"

# Use a 500-loan sample for Excel manageability
df_full = pd.read_csv(SCORED)
df = df_full.sample(500, random_state=7).reset_index(drop=True).copy()

# ---- Styles ----
BLUE_INPUT   = Font(name="Calibri", size=11, color="0000FF", bold=False)
BLACK_FORM   = Font(name="Calibri", size=11, color="000000")
GREEN_LINK   = Font(name="Calibri", size=11, color="006400")
HEADER_FONT  = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
TITLE_FONT   = Font(name="Calibri", size=16, color="1F4E79", bold=True)
SUB_FONT     = Font(name="Calibri", size=12, color="1F4E79", bold=True)
HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
INPUT_FILL   = PatternFill("solid", start_color="FFF2CC")
SUB_FILL     = PatternFill("solid", start_color="DEEBF6")
S1_FILL      = PatternFill("solid", start_color="C6EFCE")
S2_FILL      = PatternFill("solid", start_color="FFEB9C")
S3_FILL      = PatternFill("solid", start_color="FFC7CE")
THIN = Side(border_style="thin", color="999999")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR = Alignment(horizontal="center", vertical="center")
LFT = Alignment(horizontal="left", vertical="center")
RGT = Alignment(horizontal="right", vertical="center")

wb = Workbook()

# ============================================================
# SHEET 1: COVER & ASSUMPTIONS
# ============================================================
s1 = wb.active; s1.title = "Cover & Assumptions"
s1["B2"] = "UK SME Credit Risk Scorecard & IFRS 9 ECL Model"; s1["B2"].font = TITLE_FONT
s1["B3"] = "Author: Amogh H. H.   |   Methodology: PD x LGD x EAD with 3-scenario macro overlay"
s1["B3"].font = Font(italic=True, color="555555")

s1["B5"] = "Reporting date:";       s1["C5"] = "31-Mar-2026";  s1["C5"].font = BLUE_INPUT
s1["B6"] = "Reporting currency:";   s1["C6"] = "GBP";          s1["C6"].font = BLUE_INPUT
s1["B7"] = "Portfolio:";            s1["C7"] = "UK SME Lending Book"; s1["C7"].font = BLUE_INPUT

s1["B9"] = "Scenario Assumptions (IFRS 9 forward-looking)"; s1["B9"].font = SUB_FONT
s1.merge_cells("B9:F9"); s1["B9"].fill = SUB_FILL

hdr = ["Scenario","Weight","PD multiplier","LGD multiplier","UK GDP YoY"]
for j,h in enumerate(hdr):
    c = s1.cell(row=10, column=2+j, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CTR; c.border = BORDER_ALL

scns = [("Upside",0.20,0.85,0.95,0.024),
        ("Base",  0.55,1.00,1.00,0.006),
        ("Downside",0.25,1.45,1.20,-0.012)]
for i,(n,w,pm,lm,g) in enumerate(scns,start=11):
    s1.cell(row=i,column=2,value=n).font = BLACK_FORM
    s1.cell(row=i,column=3,value=w).font  = BLUE_INPUT
    s1.cell(row=i,column=4,value=pm).font = BLUE_INPUT
    s1.cell(row=i,column=5,value=lm).font = BLUE_INPUT
    s1.cell(row=i,column=6,value=g).font  = BLUE_INPUT
    s1.cell(row=i,column=3).number_format = "0.0%"
    s1.cell(row=i,column=4).number_format = "0.00"
    s1.cell(row=i,column=5).number_format = "0.00"
    s1.cell(row=i,column=6).number_format = "0.0%"
    s1.cell(row=i,column=3).fill = INPUT_FILL
    s1.cell(row=i,column=4).fill = INPUT_FILL
    s1.cell(row=i,column=5).fill = INPUT_FILL
    s1.cell(row=i,column=6).fill = INPUT_FILL
    for col in range(2,7):
        s1.cell(row=i,column=col).border = BORDER_ALL

s1["B14"] = "Total weight check:";  s1["C14"] = "=SUM(C11:C13)"
s1["C14"].number_format = "0.0%"; s1["C14"].font = BLACK_FORM

# Macro inputs
s1["B17"] = "Macro & Policy Inputs"; s1["B17"].font = SUB_FONT
s1.merge_cells("B17:F17"); s1["B17"].fill = SUB_FILL
macro = [("BoE base rate",0.0525,"0.00%"),
         ("UK unemployment rate",0.043,"0.00%"),
         ("UK CPI YoY",0.022,"0.00%"),
         ("SICR PD threshold (Stage 2 trigger)",0.15,"0.00%"),
         ("Default DPD threshold (Stage 3)",90,"0"),]
for i,(n,v,fmt) in enumerate(macro,start=18):
    s1.cell(row=i,column=2,value=n).font = BLACK_FORM
    c = s1.cell(row=i,column=3,value=v); c.font = BLUE_INPUT; c.fill = INPUT_FILL
    c.number_format = fmt; c.border = BORDER_ALL
    s1.cell(row=i,column=2).border = BORDER_ALL

s1["B25"] = "Colour key:"
s1["B26"] = "Blue text = Input (change me)"; s1["B26"].font = BLUE_INPUT
s1["B27"] = "Black text = Formula";          s1["B27"].font = BLACK_FORM
s1["B28"] = "Green text = Cross-sheet link"; s1["B28"].font = GREEN_LINK

s1.column_dimensions["A"].width = 2
s1.column_dimensions["B"].width = 38
for c in "CDEF": s1.column_dimensions[c].width = 16

# ============================================================
# SHEET 2: PORTFOLIO (LOAN-LEVEL DATA)
# ============================================================
s2 = wb.create_sheet("Portfolio")
cols = ["loan_id","sector","region","ead_gbp","lgd","days_past_due",
        "pd_calibrated","term_months","default_12m"]
headers = ["Loan ID","Sector","Region","EAD (GBP)","LGD","DPD",
           "12m PD","Term (m)","Default flag"]
for j,h in enumerate(headers,start=1):
    c = s2.cell(row=1,column=j,value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CTR; c.border = BORDER_ALL
for i,row in enumerate(df[cols].itertuples(index=False),start=2):
    for j,v in enumerate(row,start=1):
        c = s2.cell(row=i,column=j,value=v)
        c.font = BLACK_FORM; c.border = BORDER_ALL
s2.column_dimensions["A"].width = 14
s2.column_dimensions["B"].width = 22
s2.column_dimensions["C"].width = 16
for col_letter in ["D","E","F","G","H","I"]:
    s2.column_dimensions[col_letter].width = 13
# Formatting
for r in range(2, len(df)+2):
    s2.cell(row=r, column=4).number_format = '"GBP "#,##0;("GBP "#,##0);-'
    s2.cell(row=r, column=5).number_format = '0.0%'
    s2.cell(row=r, column=7).number_format = '0.00%'
s2.freeze_panes = "A2"

# ============================================================
# SHEET 3: ECL CALCULATION
# ============================================================
s3 = wb.create_sheet("ECL Calculation")
ecl_hdrs = ["Loan ID","Sector","EAD (GBP)","12m PD","LGD","DPD","Stage",
            "Horizon (m)","Lifetime PD","PD for ECL",
            "ECL Upside","ECL Base","ECL Downside","ECL Weighted"]
for j,h in enumerate(ecl_hdrs,start=1):
    c = s3.cell(row=1,column=j,value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CTR; c.border = BORDER_ALL

N = len(df)
for r in range(2, N+2):
    pr = r  # Portfolio sheet row matches
    # A: loan id from Portfolio
    s3.cell(row=r,column=1,value=f"=Portfolio!A{pr}").font = GREEN_LINK
    s3.cell(row=r,column=2,value=f"=Portfolio!B{pr}").font = GREEN_LINK
    s3.cell(row=r,column=3,value=f"=Portfolio!D{pr}").font = GREEN_LINK
    s3.cell(row=r,column=4,value=f"=Portfolio!G{pr}").font = GREEN_LINK
    s3.cell(row=r,column=5,value=f"=Portfolio!E{pr}").font = GREEN_LINK
    s3.cell(row=r,column=6,value=f"=Portfolio!F{pr}").font = GREEN_LINK
    # G Stage
    s3.cell(row=r,column=7,
        value=f'=IF(OR(F{r}>=\'Cover & Assumptions\'!$C$22,Portfolio!I{pr}=1),"Stage 3",IF(OR(F{r}>=30,D{r}>\'Cover & Assumptions\'!$C$21),"Stage 2","Stage 1"))'
    ).font = BLACK_FORM
    # H Horizon
    s3.cell(row=r,column=8,
        value=f'=IF(G{r}="Stage 1",12,MIN(Portfolio!H{pr},60))').font = BLACK_FORM
    # I Lifetime PD = 1 - (1-PD)^(h/12)
    s3.cell(row=r,column=9,
        value=f"=1-(1-D{r})^(H{r}/12)").font = BLACK_FORM
    # J PD for ECL
    s3.cell(row=r,column=10,
        value=f'=IF(G{r}="Stage 1",D{r},I{r})').font = BLACK_FORM
    # K Upside ECL = PD * LGD * EAD * pdmult * lgdmult
    s3.cell(row=r,column=11,
        value=f"=J{r}*'Cover & Assumptions'!$D$11*E{r}*'Cover & Assumptions'!$E$11*C{r}"
    ).font = BLACK_FORM
    # L Base
    s3.cell(row=r,column=12,
        value=f"=J{r}*'Cover & Assumptions'!$D$12*E{r}*'Cover & Assumptions'!$E$12*C{r}"
    ).font = BLACK_FORM
    # M Downside
    s3.cell(row=r,column=13,
        value=f"=J{r}*'Cover & Assumptions'!$D$13*E{r}*'Cover & Assumptions'!$E$13*C{r}"
    ).font = BLACK_FORM
    # N Weighted ECL
    s3.cell(row=r,column=14,
        value=f"=K{r}*'Cover & Assumptions'!$C$11+L{r}*'Cover & Assumptions'!$C$12+M{r}*'Cover & Assumptions'!$C$13"
    ).font = BLACK_FORM

# Format
for r in range(2, N+2):
    s3.cell(row=r, column=3).number_format = '"GBP "#,##0;("GBP "#,##0);-'
    s3.cell(row=r, column=4).number_format = '0.00%'
    s3.cell(row=r, column=5).number_format = '0.0%'
    s3.cell(row=r, column=9).number_format = '0.00%'
    s3.cell(row=r, column=10).number_format = '0.00%'
    for col in [11,12,13,14]:
        s3.cell(row=r, column=col).number_format = '"GBP "#,##0;("GBP "#,##0);-'
    # Conditional staging fill
    stage_cell = s3.cell(row=r, column=7)
    # Set via conditional formatting below

# Stage conditional formatting
from openpyxl.formatting.rule import FormulaRule
s3.conditional_formatting.add(f"G2:G{N+1}",
    FormulaRule(formula=['$G2="Stage 1"'], fill=S1_FILL))
s3.conditional_formatting.add(f"G2:G{N+1}",
    FormulaRule(formula=['$G2="Stage 2"'], fill=S2_FILL))
s3.conditional_formatting.add(f"G2:G{N+1}",
    FormulaRule(formula=['$G2="Stage 3"'], fill=S3_FILL))

s3.column_dimensions["A"].width = 14
s3.column_dimensions["B"].width = 22
for col_letter in "CDEFGHIJKLMN":
    s3.column_dimensions[col_letter].width = 14
s3.freeze_panes = "A2"

# ============================================================
# SHEET 4: STAGE SUMMARY
# ============================================================
s4 = wb.create_sheet("Stage Summary")
s4["B2"] = "IFRS 9 ECL - Portfolio Summary by Stage"; s4["B2"].font = TITLE_FONT
last = N+1

hdr = ["Stage","# Loans","EAD (GBP)","Weighted ECL (GBP)","Coverage Ratio"]
for j,h in enumerate(hdr,start=2):
    c = s4.cell(row=4,column=j,value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CTR; c.border = BORDER_ALL

stages = ["Stage 1","Stage 2","Stage 3"]
for i,stg in enumerate(stages,start=5):
    s4.cell(row=i,column=2,value=stg).font = BLACK_FORM
    s4.cell(row=i,column=3,value=f'=COUNTIF(\'ECL Calculation\'!G2:G{last},"{stg}")').font = BLACK_FORM
    s4.cell(row=i,column=4,value=f'=SUMIF(\'ECL Calculation\'!G2:G{last},"{stg}",\'ECL Calculation\'!C2:C{last})').font = BLACK_FORM
    s4.cell(row=i,column=5,value=f'=SUMIF(\'ECL Calculation\'!G2:G{last},"{stg}",\'ECL Calculation\'!N2:N{last})').font = BLACK_FORM
    s4.cell(row=i,column=6,value=f"=IFERROR(E{i}/D{i},0)").font = BLACK_FORM
# Totals
s4.cell(row=8,column=2,value="TOTAL").font = Font(bold=True)
s4.cell(row=8,column=3,value="=SUM(C5:C7)").font = Font(bold=True)
s4.cell(row=8,column=4,value="=SUM(D5:D7)").font = Font(bold=True)
s4.cell(row=8,column=5,value="=SUM(E5:E7)").font = Font(bold=True)
s4.cell(row=8,column=6,value="=IFERROR(E8/D8,0)").font = Font(bold=True)

for r in range(5,9):
    s4.cell(row=r,column=4).number_format = '"GBP "#,##0;("GBP "#,##0);-'
    s4.cell(row=r,column=5).number_format = '"GBP "#,##0;("GBP "#,##0);-'
    s4.cell(row=r,column=6).number_format = '0.00%'
    for c in range(2,7): s4.cell(row=r,column=c).border = BORDER_ALL

# Stage fill
s4.cell(row=5,column=2).fill = S1_FILL
s4.cell(row=6,column=2).fill = S2_FILL
s4.cell(row=7,column=2).fill = S3_FILL

# Scenario table
s4["B11"] = "ECL by Scenario (GBP)"; s4["B11"].font = SUB_FONT
s4.merge_cells("B11:F11"); s4["B11"].fill = SUB_FILL
for j,h in enumerate(["Scenario","Weight","Unweighted ECL","Weighted Contribution"], start=2):
    c = s4.cell(row=12, column=j, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CTR; c.border = BORDER_ALL
s4.cell(row=13,column=2,value="Upside")
s4.cell(row=13,column=3,value="='Cover & Assumptions'!C11")
s4.cell(row=13,column=4,value=f"=SUM('ECL Calculation'!K2:K{last})")
s4.cell(row=13,column=5,value="=C13*D13")
s4.cell(row=14,column=2,value="Base")
s4.cell(row=14,column=3,value="='Cover & Assumptions'!C12")
s4.cell(row=14,column=4,value=f"=SUM('ECL Calculation'!L2:L{last})")
s4.cell(row=14,column=5,value="=C14*D14")
s4.cell(row=15,column=2,value="Downside")
s4.cell(row=15,column=3,value="='Cover & Assumptions'!C13")
s4.cell(row=15,column=4,value=f"=SUM('ECL Calculation'!M2:M{last})")
s4.cell(row=15,column=5,value="=C15*D15")
s4.cell(row=16,column=2,value="TOTAL").font = Font(bold=True)
s4.cell(row=16,column=5,value="=SUM(E13:E15)").font = Font(bold=True)

for r in range(13,17):
    s4.cell(row=r,column=3).number_format = "0.0%"
    s4.cell(row=r,column=4).number_format = '"GBP "#,##0;("GBP "#,##0);-'
    s4.cell(row=r,column=5).number_format = '"GBP "#,##0;("GBP "#,##0);-'
    for c in range(2,6): s4.cell(row=r,column=c).border = BORDER_ALL

# Sector view
s4["B19"] = "ECL by Sector"; s4["B19"].font = SUB_FONT
s4.merge_cells("B19:F19"); s4["B19"].fill = SUB_FILL
hdrs = ["Sector","# Loans","EAD (GBP)","ECL (GBP)","Coverage %"]
for j,h in enumerate(hdrs, start=2):
    c = s4.cell(row=20,column=j,value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CTR; c.border = BORDER_ALL
sectors_sorted = sorted(df['sector'].unique())
for i,s in enumerate(sectors_sorted, start=21):
    s4.cell(row=i,column=2,value=s)
    s4.cell(row=i,column=3,value=f'=COUNTIF(\'ECL Calculation\'!B2:B{last},"{s}")')
    s4.cell(row=i,column=4,value=f'=SUMIF(\'ECL Calculation\'!B2:B{last},"{s}",\'ECL Calculation\'!C2:C{last})')
    s4.cell(row=i,column=5,value=f'=SUMIF(\'ECL Calculation\'!B2:B{last},"{s}",\'ECL Calculation\'!N2:N{last})')
    s4.cell(row=i,column=6,value=f"=IFERROR(E{i}/D{i},0)")
    s4.cell(row=i,column=4).number_format = '"GBP "#,##0;("GBP "#,##0);-'
    s4.cell(row=i,column=5).number_format = '"GBP "#,##0;("GBP "#,##0);-'
    s4.cell(row=i,column=6).number_format = '0.00%'
    for c in range(2,7): s4.cell(row=i,column=c).border = BORDER_ALL

for col_letter,w in [("A",2),("B",22),("C",14),("D",18),("E",20),("F",16)]:
    s4.column_dimensions[col_letter].width = w

# ============================================================
# SHEET 5: DASHBOARD
# ============================================================
s5 = wb.create_sheet("Dashboard"); wb.move_sheet(s5, offset=-4)
s5["B2"] = "UK SME Portfolio - IFRS 9 ECL Dashboard"; s5["B2"].font = TITLE_FONT

# KPI tiles
kpis = [
    ("Total Loans",            f"=COUNTA(Portfolio!A2:A{last})",  "0"),
    ("Total EAD",              f"=SUM(Portfolio!D2:D{last})",      '"GBP "#,##0'),
    ("Weighted ECL",           f"=SUM('ECL Calculation'!N2:N{last})", '"GBP "#,##0'),
    ("Portfolio Coverage",     f"=SUM('ECL Calculation'!N2:N{last})/SUM(Portfolio!D2:D{last})", "0.00%"),
    ("Avg 12m PD",             f"=AVERAGE(Portfolio!G2:G{last})", "0.00%"),
    ("Realised Default Rate",  f"=AVERAGE(Portfolio!I2:I{last})", "0.00%"),
    ("Stage 1 Count",          f'=COUNTIF(\'ECL Calculation\'!G2:G{last},"Stage 1")', "0"),
    ("Stage 2 Count",          f'=COUNTIF(\'ECL Calculation\'!G2:G{last},"Stage 2")', "0"),
    ("Stage 3 Count",          f'=COUNTIF(\'ECL Calculation\'!G2:G{last},"Stage 3")', "0"),
]
for i,(label,formula,fmt) in enumerate(kpis):
    row = 5 + (i//3)*3
    col = 2 + (i%3)*3
    lc = s5.cell(row=row, column=col, value=label)
    lc.font = HEADER_FONT; lc.fill = HEADER_FILL; lc.alignment = CTR
    s5.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+1)
    vc = s5.cell(row=row+1, column=col, value=formula)
    vc.font = Font(name="Calibri",size=18,bold=True,color="1F4E79"); vc.alignment = CTR
    vc.number_format = fmt
    s5.merge_cells(start_row=row+1,start_column=col,end_row=row+1,end_column=col+1)
    for c in range(col, col+2):
        s5.cell(row=row,column=c).border = BORDER_ALL
        s5.cell(row=row+1,column=c).border = BORDER_ALL

# Notes
s5["B17"] = "Notes:"; s5["B17"].font = SUB_FONT
s5["B18"] = "- All values are model outputs (formulas). Edit scenario weights on 'Cover & Assumptions' to flex."
s5["B19"] = "- Stage 2 trigger: PD > SICR threshold or DPD >= 30."
s5["B20"] = "- Stage 3 trigger: DPD >= 90 or default flag = 1."
s5["B21"] = "- ECL = PD x LGD x EAD with probability-weighted 3-scenario macro overlay (IFRS 9 compliant)."

for col_letter in "ABCDEFGHIJ":
    s5.column_dimensions[col_letter].width = 14

wb.save(OUT_XLSX)
print(f"Saved {OUT_XLSX}")
print(f"Sheets: {wb.sheetnames}")
